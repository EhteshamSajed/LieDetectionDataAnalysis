from os import listdir
from openpyxl import Workbook
import json
import Utilities

filePath = "ExcelExport/"

pupilSizeColumn = {
    1: "B",
    2: "C",
    4: "D",
    5: "E"
}

responseTimeColumn = {
    1: "G",
    2: "H",
    4: "I",
    5: "J"
}

numberColumn = {
    1: "L",
    2: "M",
    4: "N",
    5: "O"
}

fileNameDictionary = {
    0: "AllAverage_NoFeedback.xlsx",
    1: "AllAverage_Feedback.xlsx"
}

def getAveragePupilSize(feedbackCondition):
    fileName = fileNameDictionary[feedbackCondition]
    wb = Workbook()
    ws = wb.active
    ws['A1'].value = "Participants"
    ws['B1'].value = "Trues"
    ws['C1'].value = "Lies"
    ws['D1'].value = "Free Trues"
    ws['E1'].value = "Free Lies"
    
    ws['G1'].value = "True Sec."
    ws['H1'].value = "Lies Sec."
    ws['I1'].value = "Free Trues Sec."
    ws['J1'].value = "Free Lies Sec."

    ws['L1'].value = "# Trues"
    ws['M1'].value = "# Lies"
    ws['N1'].value = "# Free Trues"
    ws['O1'].value = "# Free Lies"

    dir = "ExpData/V2/"
    experiment_files = listdir(dir)
    scope = Utilities.Trial_Data.baseline_difference
    # scope = Utilities.Trial_Data.baseline_difference_decision_phase
    # feedbackCondition = 0
    # condition_index = 2
    search_from = 0
    count = 30
    conditions = [2, 1, 4, 5]
    # conditions = [0, 2, 1]
    # conditions = [4, 5]
    # condition = Utilities.CONDITIONS[condition_index]
    i = 2
    for file in experiment_files:
        data = json.load(open(dir + file))
        for condition_index in conditions:
            condition = Utilities.CONDITIONS[condition_index]
            # data = json.load(open(dir + file))
            extracted = Utilities.extract_data(
                data=data, search_from=search_from, count=count, feedbackCondition=feedbackCondition, participantAnswer=0,
                    condition_index=condition_index, baseline_source=Utilities.Baseline_Source.preceding_trial
            )
            if len(extracted) == 0:
                continue
            average_within_condition = Utilities.average_within_condition(
                    extracted, scope.name, condition)     
            ws[pupilSizeColumn[condition_index] + str(i)].value = average_within_condition["average_pupil_size"]
            ws[responseTimeColumn[condition_index] + str(i)].value = average_within_condition["average_elapsed_ticks_to_answer"] / 10000000
            ws[numberColumn[condition_index] + str(i)].value = average_within_condition["number_of_data"]

        ws['A' + str(i)].value = file.removesuffix(".dat")
        i += 1

    wb.save(filePath+fileName)
    print ("Exported to file: "+ filePath+fileName)
    

getAveragePupilSize(0)
getAveragePupilSize(1)